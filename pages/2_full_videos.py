import os
import re

from googleapiclient.discovery import build
from googletrans import Translator
import requests

import streamlit as st

from PIL import Image as PILImage
import pandas as pd

from colorstyle import color_style

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill


API_KEY = "AIzaSyD_NcigzVQdRDH5EZcaK7bdrw01hDXfrFw"
file_path = 'output.xlsx'
excel_pattern = 'pattern.xlsx'

translator = Translator()


def get_service():
    service = build('youtube', 'v3', developerKey=API_KEY)
    return service


def get_channel_subs_count(channel_id):
    r = get_service().channels().list(id=channel_id, part='snippet,statistics').execute()
    return r['items'][0]['statistics']['subscriberCount']


# Функция для добавления новой строки в Excel файл
def append_to_excel(file_path, data):
    if not os.path.isfile(file_path):
        # Загружаем оригинальный файл
        wb = load_workbook(excel_pattern)
        # Сохраняем его под новым именем
        wb.save(file_path)

    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Проверяем, есть ли видео с таким URL
    video_urls = [sheet.cell(row=i, column=2).value for i in range(2, sheet.max_row + 1)]
    if data[0] in video_urls:
        st.warning("Видео с таким URL уже существует в таблице!")
        return

    max_rows = sheet.max_row
    target_row = max_rows + 1
    cell_value_id = sheet.cell(row=max_rows, column=1).value

    sheet.row_dimensions[target_row].height = 135

    sheet.cell(row=max_rows + 1, column=1, value=int(cell_value_id) + 1)
    sheet.cell(row=max_rows + 1, column=2, value=data[0])
    sheet.cell(row=max_rows + 1, column=3, value=data[1])

    # Вставьте изображение
    img = Image(data[2])
    img.anchor = f"D{target_row}"  # Установите анкор изображения в D и соответствующую строку
    sheet.add_image(img)

    sheet.cell(row=max_rows + 1, column=5, value=data[3])
    sheet.cell(row=max_rows + 1, column=6, value=data[4])
    sheet.cell(row=max_rows + 1, column=7, value=data[5])
    sheet.cell(row=max_rows + 1, column=8, value=data[6])

    sheet.cell(row=max_rows + 1, column=9, value=data[7])
    sheet[f'I{target_row}'].fill = PatternFill(
        start_color=color_style[data[7].split('-')[0]],
        end_color=color_style[data[7].split('-')[0]],
        fill_type='solid'
    )

    try:
        workbook.save(file_path)
    except Exception as ex:
        st.error(f"Возникла ошибка при сохранении таблицы: {ex}")
        return

    st.success(f"Видео добавлено в таблицу! Всего строк в таблице: {max_rows - 1}")


def load_table_info(file_path):
    # Загружаем данные из Excel
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active

    # Список для хранения данных
    data = []

    # Читаем строки из таблицы, начиная со 2-й строки (первая — заголовки)
    for i, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=8), start=1):
        title = row[2].value  # 3-й столбец: название видео
        url = row[1].value  # 4-й столбец: ссылка
        ratio = row[7].value  # 8-й столбец: просмотры/подписчики
        data.append({"Номер": i, "Название": title, "Ссылка": url, "Просмотры/Подписчики": ratio})

    # Преобразуем список в DataFrame
    df = pd.DataFrame(data)

    # Выводим таблицу в Streamlit
    st.write("### Таблица аналитики")
    for index, row in df.iterrows():
        # Создаём строку с колонками
        cols = st.columns([1, 3, 3, 2, 1])

        with cols[0]:
            st.write(row["Номер"])  # Автоматически пронумерованная строка

        with cols[1]:
            st.write(row["Название"] if row["Название"] else "Нет названия")  # Название видео

        with cols[2]:
            # Добавляем кликабельную ссылку
            if row["Ссылка"]:
                st.markdown(f"[Смотреть видео]({row['Ссылка']})")
            else:
                st.write("Нет ссылки")

        with cols[3]:
            st.write(row["Просмотры/Подписчики"] if row["Просмотры/Подписчики"] else "Нет данных")

        with cols[4]:
            # Кнопка для удаления строки
            if st.button(f"Удалить {row['Номер']}", key=f"delete_{row['Номер']}"):
                try:
                    delete_row(file_path, index + 2)
                    st.query_params.clear()  # Очищаем параметры URL
                    st.rerun()
                except PermissionError:
                    vote()


# Функция для удаления строки из Excel
def delete_row(file_path, row_number):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    # Удаляем строку
    sheet.delete_rows(row_number)
    # Сохраняем изменения
    workbook.save(file_path)


@st.dialog("Возникла ошибка")
def vote():
    st.write("Возможно, у вас открыл Excel файл. Закройте его и повторите попытку")
    if st.button("Ок"):
        st.rerun()


def get_video_info(video_id):
    try:
        # Получение данных о видео
        r = get_service().videos().list(id=video_id, part='snippet, statistics, contentDetails').execute()

        channelID = r['items'][0]['snippet']['channelId']
        channelSubsCount = get_channel_subs_count(channelID)

        # Получение названия видео
        videoTitle = r['items'][0]['snippet'].get('title', 'Без названия')
        if not videoTitle:  # Если название пустое или отсутствует
            videoTitle = "Без названия"

        # Перевод названия видео на русский язык
        try:
            translated_title = translator.translate(videoTitle, src='auto', dest='ru').text
        except Exception as e:
            translated_title = videoTitle  # Если перевод не удался, оставляем оригинал
            print(f"Ошибка при переводе: {e}")

        videoDatePublish = r['items'][0]['snippet']['publishedAt']
        videoViewsCount = r['items'][0]['statistics']['viewCount']

        videoDuration = r['items'][0]['contentDetails']['duration']
        duration = videoDuration.replace("PT", "").replace("S", "")
        duration_on_write = str(duration.split('M')[0] + ":" + duration.split('M')[1])

        urlPreview = r['items'][0]['snippet']['thumbnails']['medium']['url']
        p = requests.get(urlPreview)
        previewPath = f"previews\{video_id}.jpg"
        with open(previewPath, "wb") as out:
            out.write(p.content)

        # Сохраняем перевод в таблицу
        videos_data = [f"https://www.youtube.com/watch?v={video_id}",
                       translated_title,  # Используем переведённое название
                       previewPath,
                       videoViewsCount,
                       channelSubsCount,
                       duration_on_write,
                       round(int(videoViewsCount) / int(channelSubsCount), 2),
                       videoDatePublish.split('T')[0]]

        append_to_excel(file_path, videos_data)

    except KeyError as e:
        st.error(f"Ошибка: отсутствует ключ {e}. Возможно, видео недоступно.")
    except IndexError:
        pass
    except Exception as e:
        st.error(f"Возникла ошибка: {e}")


def get_youtube_id(url):
    # Паттерн для поиска ID видео
    pattern = r"(?:youtu\.be\/|(?:www\.)?youtube\.com\/(?:watch\?v=|embed\/|v\/|.+\?v=))([^&]{11})"
    match = re.search(pattern, url)

    if match:
        return match.group(1)
    else:
        return None


def main():
    # Проверка авторизации
    if "authenticated" not in st.session_state or not st.session_state["authenticated"]:
        st.error("Доступ запрещен. Пожалуйста, войдите в систему на главной странице.")
        return

    st.page_link("main.py", label="На главную", icon="⬅️")
    st.title("АНАЛИТИКА—ХУИТИКА | Полноформатные видео")

    yt_link = st.text_input("Ссылка на ютуб ролик")

    #  correct_link = str(yt_link.split('=')[-1])
    correct_link = ""
    if yt_link != "":
        correct_link = get_youtube_id(yt_link)

    try:
        get_video_info(correct_link)
    except IndexError:
        if correct_link != "":
            st.error(f"Возникла ошибка: Неккоректно указана ссылка!")
    except PermissionError:
        st.error(f"Возникла ошибка: Возможно, у вас открыл Excel файл. Закройте его и повторите попытку")
    except Exception as ex:
        st.error(f"Возникла ошибка: {ex}")

    if os.path.isfile(file_path):
        pass
        load_table_info(file_path)


if __name__ == "__main__":
    main()
